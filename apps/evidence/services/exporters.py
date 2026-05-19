from __future__ import annotations

import html
import json
import zipfile
from datetime import datetime
from pathlib import Path
from typing import Any

from .collectors import EvidenceSource, write_redacted_json
from .redaction import redact_obj, redact_text

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import LETTER, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
except Exception:  # pragma: no cover
    colors = None
    LETTER = None
    landscape = None
    getSampleStyleSheet = None
    Paragraph = None
    SimpleDocTemplate = None
    Spacer = None
    Table = None
    TableStyle = None

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
except Exception:  # pragma: no cover
    Workbook = None


def _findings_list(findings: Any) -> list[dict[str, Any]]:
    if isinstance(findings, dict):
        for key in ("findings", "items", "results"):
            if isinstance(findings.get(key), list):
                return [x if isinstance(x, dict) else {"finding": x} for x in findings[key]]
        # Support switch keyed format
        rows: list[dict[str, Any]] = []
        for switch, value in findings.items():
            if isinstance(value, list):
                for item in value:
                    row = item if isinstance(item, dict) else {"finding": item}
                    row.setdefault("switch", switch)
                    rows.append(row)
        return rows
    if isinstance(findings, list):
        return [x if isinstance(x, dict) else {"finding": x} for x in findings]
    return []


def _port_rows(port_map: Any) -> list[dict[str, Any]]:
    if isinstance(port_map, dict):
        for key in ("ports", "rows", "port_map"):
            if isinstance(port_map.get(key), list):
                return [x if isinstance(x, dict) else {"port": x} for x in port_map[key]]
        rows: list[dict[str, Any]] = []
        for switch, value in port_map.items():
            if isinstance(value, list):
                for item in value:
                    row = item if isinstance(item, dict) else {"port": item}
                    row.setdefault("switch", switch)
                    rows.append(row)
        return rows
    if isinstance(port_map, list):
        return [x if isinstance(x, dict) else {"port": x} for x in port_map]
    return []


def _severity_counts(rows: list[dict[str, Any]]) -> dict[str, int]:
    counts = {"critical": 0, "warning": 0, "info": 0, "other": 0}
    for row in rows:
        sev = str(row.get("severity") or row.get("level") or "other").lower()
        if sev not in counts:
            sev = "other"
        counts[sev] += 1
    return counts


def build_offline_html(source: EvidenceSource) -> Path:
    findings = _findings_list(source.findings)
    counts = _severity_counts(findings)
    port_count = len(_port_rows(source.port_map))
    raw_count = len(source.raw_files)
    generated = datetime.now().isoformat(timespec="seconds")

    finding_rows = "".join(
        f"<tr><td>{html.escape(str(f.get('severity') or f.get('level') or ''))}</td>"
        f"<td>{html.escape(str(f.get('switch') or f.get('device') or f.get('host') or ''))}</td>"
        f"<td>{html.escape(str(f.get('title') or f.get('finding' ) or f.get('message') or f))}</td>"
        f"<td>{html.escape(str(f.get('remediation') or f.get('suggestion') or ''))}</td></tr>"
        for f in findings[:500]
    )
    warnings = "".join(f"<li>{html.escape(w)}</li>" for w in source.warnings)

    raw_links = "".join(
        f"<li><a href='raw_cli_output/{html.escape(p.name)}'>{html.escape(p.name)}</a></li>" for p in source.raw_files
    )

    html_doc = f"""<!doctype html>
<html lang="en">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Single Digits Evidence Pack - {html.escape(source.session_id)}</title>
<style>
:root {{ --blue:#14345f; --cyan:#00a6d6; --bg:#f4f7fb; --card:#fff; --text:#172033; --muted:#65728a; --bad:#a32020; --warn:#9a6400; }}
* {{ box-sizing:border-box; }} body {{ margin:0; font-family:Segoe UI, Arial, sans-serif; background:var(--bg); color:var(--text); }}
header {{ background:linear-gradient(135deg,var(--blue),#0b203d); color:white; padding:28px 36px; }}
header h1 {{ margin:0 0 6px; font-size:28px; }} header p {{ margin:0; color:#dce9f7; }}
main {{ padding:24px 36px 48px; }} .grid {{ display:grid; grid-template-columns:repeat(auto-fit,minmax(180px,1fr)); gap:16px; margin-bottom:22px; }}
.card {{ background:var(--card); border-radius:16px; padding:18px; box-shadow:0 8px 24px rgba(20,52,95,.08); border:1px solid #e5edf7; }}
.metric {{ font-size:30px; font-weight:700; color:var(--blue); }} .label {{ color:var(--muted); font-size:13px; text-transform:uppercase; letter-spacing:.06em; }}
section {{ margin-top:22px; }} h2 {{ color:var(--blue); margin:0 0 12px; }} table {{ width:100%; border-collapse:collapse; background:white; border-radius:12px; overflow:hidden; }}
th,td {{ padding:10px 12px; border-bottom:1px solid #e8eef6; text-align:left; vertical-align:top; }} th {{ background:#eaf2fb; color:#14345f; }}
.badge {{ display:inline-block; padding:4px 8px; border-radius:999px; background:#eaf2fb; color:#14345f; font-size:12px; }}
footer {{ color:var(--muted); padding:16px 36px 36px; font-size:12px; }}
</style>
</head>
<body>
<header><h1>Single Digits Evidence Pack</h1><p>Session {html.escape(source.session_id)} • Generated {generated} • Offline self-contained report</p></header>
<main>
<div class="grid">
  <div class="card"><div class="label">Critical</div><div class="metric">{counts['critical']}</div></div>
  <div class="card"><div class="label">Warnings</div><div class="metric">{counts['warning']}</div></div>
  <div class="card"><div class="label">Info</div><div class="metric">{counts['info']}</div></div>
  <div class="card"><div class="label">Port Rows</div><div class="metric">{port_count}</div></div>
  <div class="card"><div class="label">Raw CLI Files</div><div class="metric">{raw_count}</div></div>
</div>
<section class="card"><h2>Executive Summary</h2><p>This package captures scan evidence, findings, topology, port mapping, raw command output, and session logs from the Single Digits Engineering Platform. Credentials, SNMP strings, tokens, and shared secrets are redacted before export.</p></section>
<section class="card"><h2>Findings</h2><table><thead><tr><th>Severity</th><th>Switch</th><th>Finding</th><th>Remediation</th></tr></thead><tbody>{finding_rows or '<tr><td colspan="4">No findings were present in the collected session data.</td></tr>'}</tbody></table></section>
<section class="card"><h2>Included Raw CLI Output</h2><ul>{raw_links or '<li>No raw CLI files found.</li>'}</ul></section>
<section class="card"><h2>Collection Warnings</h2><ul>{warnings or '<li>No collection warnings.</li>'}</ul></section>
</main><footer>Single Digits Engineering Platform • Evidence Pack Module • Sensitive values redacted</footer>
</body></html>"""
    out = source.work_dir / "session_report.html"
    out.write_text(redact_text(html_doc), encoding="utf-8")
    return out


def _write_simple_pdf(path: Path, title: str, source: EvidenceSource, kind: str = "executive") -> Path:
    if SimpleDocTemplate is None:
        path.write_text(f"{title}\n\nReportLab is not installed. Install reportlab to generate PDF output.", encoding="utf-8")
        return path
    doc = SimpleDocTemplate(str(path), pagesize=LETTER, rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=36)
    styles = getSampleStyleSheet()
    story = [Paragraph(title, styles["Title"]), Spacer(1, 10)]
    story.append(Paragraph(f"Session: {source.session_id}", styles["Normal"]))
    story.append(Paragraph(f"Generated: {datetime.now().isoformat(timespec='seconds')}", styles["Normal"]))
    story.append(Spacer(1, 14))

    if kind == "executive":
        findings = _findings_list(source.findings)
        counts = _severity_counts(findings)
        summary = (
            "This evidence package summarizes the scan session in a leadership and vendor-friendly format. "
            "Sensitive credentials and SNMP/community values are redacted before export."
        )
        story.append(Paragraph(summary, styles["BodyText"]))
        story.append(Spacer(1, 12))
        table_data = [["Critical", "Warning", "Info", "Other", "Raw CLI Files"], [counts["critical"], counts["warning"], counts["info"], counts["other"], len(source.raw_files)]]
        tbl = Table(table_data, hAlign="LEFT")
        tbl.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#14345f")), ("TEXTCOLOR", (0, 0), (-1, 0), colors.white), ("GRID", (0, 0), (-1, -1), .25, colors.grey), ("PADDING", (0, 0), (-1, -1), 8)]))
        story.append(tbl)
        story.append(Spacer(1, 14))
        top_findings = findings[:20]
        if top_findings:
            rows = [["Severity", "Switch", "Finding"]]
            for f in top_findings:
                rows.append([redact_text(f.get("severity") or f.get("level") or ""), redact_text(f.get("switch") or f.get("device") or ""), redact_text(f.get("title") or f.get("finding") or f.get("message") or str(f))[:180]])
            tbl = Table(rows, colWidths=[70, 110, 330])
            tbl.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#eaf2fb")), ("GRID", (0, 0), (-1, -1), .25, colors.lightgrey), ("VALIGN", (0, 0), (-1, -1), "TOP")]))
            story.append(tbl)
    else:
        payload = source.topology if kind == "topology" else source.port_map
        story.append(Paragraph(redact_text(json.dumps(redact_obj(payload), indent=2)[:6000]).replace("\n", "<br/>") or "No data found.", styles["Code"]))

    doc.build(story)
    return path


def build_pdfs(source: EvidenceSource) -> dict[str, Path]:
    return {
        "executive_summary.pdf": _write_simple_pdf(source.work_dir / "executive_summary.pdf", "Executive Summary", source, "executive"),
        "topology.pdf": _write_simple_pdf(source.work_dir / "topology.pdf", "Topology Evidence", source, "topology"),
        "port_map.pdf": _write_simple_pdf(source.work_dir / "port_map.pdf", "Port Map Evidence", source, "port_map"),
    }


def build_excel_port_map(source: EvidenceSource) -> Path:
    out = source.work_dir / "excel_port_map_export.xlsx"
    if Workbook is None:
        out.write_text("openpyxl is not installed. Install openpyxl to generate Excel output.", encoding="utf-8")
        return out
    wb = Workbook()
    ws = wb.active
    ws.title = "Port Map"
    rows = _port_rows(source.port_map)
    headers = ["Switch", "Port", "Status", "VLAN", "Tagged VLANs", "Device", "LLDP Neighbor", "MAC Count", "Notes", "Patch Panel"]
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="14345F")
        c.alignment = Alignment(horizontal="center")
    for row in rows:
        ws.append([
            redact_text(row.get("switch") or row.get("hostname") or row.get("device_name") or ""),
            redact_text(row.get("port") or row.get("interface") or ""),
            redact_text(row.get("status") or row.get("link") or ""),
            redact_text(row.get("vlan") or row.get("untagged_vlan") or ""),
            redact_text(", ".join(map(str, row.get("tagged_vlans", []))) if isinstance(row.get("tagged_vlans"), list) else row.get("tagged_vlans") or ""),
            redact_text(row.get("device") or ""),
            redact_text(row.get("lldp_neighbor") or row.get("neighbor") or ""),
            redact_text(row.get("mac_count") or ""),
            redact_text(row.get("notes") or ""),
            "",  # Patch Panel must remain blank
        ])
    for col in ws.columns:
        width = max(12, min(42, max(len(str(cell.value or "")) for cell in col) + 2))
        ws.column_dimensions[col[0].column_letter].width = width
    wb.save(out)
    return out


def build_zip(source: EvidenceSource) -> Path:
    build_offline_html(source)
    build_pdfs(source)
    build_excel_port_map(source)
    write_redacted_json(source.work_dir / "findings.json", source.findings)
    write_redacted_json(source.work_dir / "topology.json", source.topology)
    write_redacted_json(source.work_dir / "port_map.json", source.port_map)
    (source.work_dir / "session.log").write_text(redact_text(source.session_log), encoding="utf-8")

    manifest = {
        "session_id": source.session_id,
        "generated": datetime.now().isoformat(timespec="seconds"),
        "source_dir": str(source.source_dir),
        "redaction": "credentials, SNMP strings, community strings, tokens, and secrets redacted",
        "warnings": source.warnings,
    }
    write_redacted_json(source.work_dir / "manifest.json", manifest)

    zip_path = source.work_dir.with_suffix(".zip")
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
        for path in source.work_dir.rglob("*"):
            if path.is_file():
                zf.write(path, path.relative_to(source.work_dir))
    return zip_path
