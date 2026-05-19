from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

EXPORT_DIR = Path(__file__).resolve().parent / "exports"
EXPORT_DIR.mkdir(parents=True, exist_ok=True)
TEMPLATE_PATH = Path(__file__).resolve().parent / "Switch Template.xlsx"

PORT_HEADERS = [
    "Switch", "NAT IP", "Vendor", "Port", "Patch Panel", "Device", "Description",
    "Status", "Speed", "Duplex", "VLAN", "Tagged VLANs", "MAC Count", "Category", "Rename Preview"
]


def export_json(job: Dict[str, Any]) -> Path:
    path = EXPORT_DIR / f"port_map_{job.get('job_id', 'export')}.json"
    path.write_text(json.dumps(job, indent=2), encoding="utf-8")
    return path


def export_excel(job: Dict[str, Any]) -> Path:
    if TEMPLATE_PATH.exists():
        wb = load_workbook(TEMPLATE_PATH)
        ws = wb.active
        ws.title = ws.title or "Port Map"
        clear_body(ws)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Port Map"
        ws.append(PORT_HEADERS)
    ensure_headers(ws)
    raw = wb.create_sheet("Raw Data") if "Raw Data" not in wb.sheetnames else wb["Raw Data"]
    raw.delete_rows(1, raw.max_row)
    raw.append(["Switch", "IP", "Vendor", "Port", "Field", "Value"])

    for sw in job.get("switches", []):
        for p in sw.get("ports", []):
            lldp = p.get("lldp") or {}
            # Patch Panel column is intentionally and always blank.
            # Device only receives high-confidence LLDP system-name/hostname.
            ws.append([
                sw.get("hostname", ""), sw.get("ip", ""), sw.get("vendor", ""), p.get("port", ""), "",
                p.get("confident_device", ""), p.get("description", ""), p.get("status", ""), p.get("speed", ""),
                p.get("duplex", ""), p.get("vlan", ""), ", ".join(p.get("tagged_vlans", [])),
                p.get("mac_count", 0), p.get("category", ""), p.get("rename_suggestion", ""),
            ])
            raw.append([sw.get("hostname", ""), sw.get("ip", ""), sw.get("vendor", ""), p.get("port", ""), "lldp_raw", lldp.get("raw", "")])
            raw.append([sw.get("hostname", ""), sw.get("ip", ""), sw.get("vendor", ""), p.get("port", ""), "raw", json.dumps(p.get("raw", {}))])
            for mac in p.get("macs", []):
                raw.append([sw.get("hostname", ""), sw.get("ip", ""), sw.get("vendor", ""), p.get("port", ""), "mac", json.dumps(mac)])
    style_sheet(ws)
    style_sheet(raw)
    path = EXPORT_DIR / f"port_map_{job.get('job_id', 'export')}.xlsx"
    wb.save(path)
    return path


def clear_body(ws):
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row)


def ensure_headers(ws):
    existing = [ws.cell(1, c).value for c in range(1, len(PORT_HEADERS) + 1)]
    if existing != PORT_HEADERS:
        for c, h in enumerate(PORT_HEADERS, 1):
            ws.cell(1, c).value = h


def style_sheet(ws):
    fill = PatternFill("solid", fgColor="1F2937")
    font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"
    for col in range(1, min(ws.max_column, 15) + 1):
        letter = get_column_letter(col)
        ws.column_dimensions[letter].width = 18 if col not in (6, 15) else 28
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")


def export_pdf(job: Dict[str, Any]) -> Path:
    path = EXPORT_DIR / f"port_map_executive_{job.get('job_id', 'export')}.pdf"
    doc = SimpleDocTemplate(str(path), pagesize=landscape(letter), rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    styles = getSampleStyleSheet()
    story = [Paragraph("Single Digits Port Map Executive Summary", styles["Title"]), Paragraph(datetime.now().strftime("Generated %Y-%m-%d %H:%M"), styles["Normal"]), Spacer(1, 12)]
    switches = job.get("switches", [])
    ports = [p for sw in switches for p in sw.get("ports", [])]
    infra = [p for p in ports if p.get("category") == "infrastructure"]
    aps = [p for p in ports if p.get("category") == "ap"]
    edge = [p for p in ports if p.get("category") in ("edge", "endpoint")]
    summary = [["Switches", "Ports", "Infrastructure Links", "AP Links", "Edge/Endpoint Links"], [len(switches), len(ports), len(infra), len(aps), len(edge)]]
    story.append(Table(summary, style=base_table_style()))
    story.append(Spacer(1, 14))
    rows = [["Switch", "IP", "Vendor", "Port", "Infrastructure Neighbor", "Speed", "Status"]]
    for sw in switches:
        for p in sw.get("ports", []):
            if p.get("category") == "infrastructure":
                rows.append([sw.get("hostname", ""), sw.get("ip", ""), sw.get("vendor", ""), p.get("port", ""), p.get("confident_device") or (p.get("lldp") or {}).get("display_name", ""), p.get("speed", ""), p.get("status", "")])
    if len(rows) == 1:
        rows.append(["No infrastructure LLDP links found", "", "", "", "", "", ""])
    story.append(Paragraph("Prioritized Infrastructure Links", styles["Heading2"]))
    story.append(Table(rows, repeatRows=1, style=base_table_style()))
    doc.build(story)
    return path


def base_table_style():
    return TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F2937")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F3F4F6")]),
    ])
